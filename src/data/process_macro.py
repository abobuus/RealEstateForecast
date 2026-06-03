"""
Обработка макроэкономических данных → CSV файлы в data/macro/processed/
"""
import pandas as pd
import numpy as np
import openpyxl
from docx import Document
import os, re

OUT_DIR = 'data/macro/processed'
os.makedirs(OUT_DIR, exist_ok=True)


# ─────────────────────────────────────────────
# 1. Ключевая ставка + инфляция
# ─────────────────────────────────────────────
def process_cbr_inflation():
    wb = openpyxl.load_workbook(
        'data/macro/Инфляция и ключевая ставка Банка России_F01_01_2018_T31_12_2021(1).xlsx'
    )
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    records = []
    for row in rows:
        if row[0] is None:
            continue
        date_str = str(row[0])  # формат MM.YYYY
        month, year = date_str.split('.')
        date = f'{year}-{month}'
        cbr_rate   = row[1]  # ключевая ставка, %
        inflation  = row[2]  # инфляция г/г, %
        records.append({'date': date, 'cbr_rate': cbr_rate, 'inflation': inflation})

    df = pd.DataFrame(records).sort_values('date').reset_index(drop=True)
    df.to_csv(f'{OUT_DIR}/cbr_rate.csv', index=False)
    df[['date', 'inflation']].to_csv(f'{OUT_DIR}/inflation.csv', index=False)
    print(f'cbr_rate.csv: {len(df)} строк, {df["date"].min()} — {df["date"].max()}')
    print(f'inflation.csv: {len(df)} строк')


# ─────────────────────────────────────────────
# 2. Курс доллара (ежедневный → помесячный)
# ─────────────────────────────────────────────
def process_exchange_rate():
    wb = openpyxl.load_workbook('data/macro/RC_F01_01_2018_T31_12_2021.xlsx')
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    records = []
    for row in rows:
        if row[1] is None:
            continue
        date = pd.to_datetime(row[1])
        rate = row[2]
        records.append({'date': date, 'usd_rate': rate})

    df = pd.DataFrame(records)
    df['date'] = pd.to_datetime(df['date'])
    # Агрегируем по месяцам (среднее за месяц)
    df['month'] = df['date'].dt.to_period('M').astype(str)
    monthly = df.groupby('month')['usd_rate'].mean().round(2).reset_index()
    monthly.columns = ['date', 'usd_rate']
    monthly = monthly.sort_values('date').reset_index(drop=True)
    monthly.to_csv(f'{OUT_DIR}/exchange_rate.csv', index=False)
    print(f'exchange_rate.csv: {len(monthly)} строк, {monthly["date"].min()} — {monthly["date"].max()}')


# ─────────────────────────────────────────────
# 3. Доходы населения по регионам
# ─────────────────────────────────────────────

# Маппинг названий регионов из Росстата → официальный код
REGION_NAME_TO_CODE = {
    'Республика Адыгея': 1, 'Республика Башкортостан': 2, 'Республика Бурятия': 3,
    'Республика Алтай': 4, 'Республика Дагестан': 5, 'Республика Ингушетия': 6,
    'Кабардино-Балкарская Республика': 7, 'Республика Калмыкия': 8,
    'Карачаево-Черкесская Республика': 9, 'Республика Карелия': 10,
    'Республика Коми': 11, 'Республика Марий Эл': 12, 'Республика Мордовия': 13,
    'Республика Саха (Якутия)': 14, 'Республика Северная Осетия - Алания': 15,
    'Республика Татарстан': 16, 'Республика Тыва': 17, 'Удмуртская Республика': 18,
    'Республика Хакасия': 19, 'Чеченская Республика': 20, 'Чувашская Республика': 21,
    'Алтайский край': 22, 'Краснодарский край': 23, 'Красноярский край': 24,
    'Приморский край': 25, 'Ставропольский край': 26, 'Хабаровский край': 27,
    'Амурская область': 28, 'Архангельская область': 29, 'Астраханская область': 30,
    'Белгородская область': 31, 'Брянская область': 32, 'Владимирская область': 33,
    'Волгоградская область': 34, 'Вологодская область': 35, 'Воронежская область': 36,
    'Ивановская область': 37, 'Иркутская область': 38, 'Калининградская область': 39,
    'Калужская область': 40, 'Камчатский край': 41, 'Кемеровская область': 42,
    'Кировская область': 43, 'Костромская область': 44, 'Курганская область': 45,
    'Курская область': 46, 'Ленинградская область': 47, 'Липецкая область': 48,
    'Магаданская область': 49, 'Московская область': 50, 'Мурманская область': 51,
    'Нижегородская область': 52, 'Новгородская область': 53, 'Новосибирская область': 54,
    'Омская область': 55, 'Оренбургская область': 56, 'Орловская область': 57,
    'Пензенская область': 58, 'Пермский край': 59, 'Псковская область': 60,
    'Ростовская область': 61, 'Рязанская область': 62, 'Самарская область': 63,
    'Саратовская область': 64, 'Сахалинская область': 65, 'Свердловская область': 66,
    'Смоленская область': 67, 'Тамбовская область': 68, 'Тверская область': 69,
    'Томская область': 70, 'Тульская область': 71, 'Тюменская область': 72,
    'Ульяновская область': 73, 'Челябинская область': 74, 'Забайкальский край': 75,
    'Ярославская область': 76, 'г. Москва': 77, 'г. Санкт-Петербург': 78,
    'Еврейская автономная область': 79, 'Ненецкий автономный округ': 83,
    'Ханты-Мансийский автономный округ - Югра': 86,
    'Чукотский автономный округ': 87, 'Ямало-Ненецкий автономный округ': 89,
    'Республика Крым': 91, 'г. Севастополь': 92,
    # Альтернативные написания
    'Кемеровская область - Кузбасс': 42,
    'в том числе Ханты-Мансийский автономный округ - Ю': 86,
    'Ямало-Ненецкий автономный округ без автономного о': 89,
    'Тюменская область без автономных округов': 72,
}

def clean_value(s):
    """Очищаем строку с числом типа '58 781 490' → float"""
    s = s.replace('\xa0', '').replace(' ', '').replace(',', '.')
    try:
        return float(s)
    except:
        return None

def process_income():
    doc = Document('data/macro/den-dohod 18-21/razd1-Таблица 1.docx')
    t = doc.tables[0]

    YEARS = [2018, 2019, 2020, 2021]
    records = []
    unmatched = []

    for ri, row in enumerate(t.rows):
        name = row.cells[0].text.strip().replace('\n', ' ').strip()
        if not name or ri < 2:
            continue

        # Пробуем найти код региона
        code = None
        for key, val in REGION_NAME_TO_CODE.items():
            if name.startswith(key) or key.startswith(name[:20]):
                code = val
                break
        if code is None:
            # Попытка частичного совпадения
            for key, val in REGION_NAME_TO_CODE.items():
                if name[:15] in key or key[:15] in name:
                    code = val
                    break

        if code is None:
            unmatched.append(name)
            continue

        # Извлекаем значения за каждый год
        for yi, year in enumerate(YEARS):
            val_str = row.cells[yi + 1].text.strip()
            val = clean_value(val_str)
            if val is not None:
                records.append({'year': year, 'region': code, 'income_mln': val})

    df = pd.DataFrame(records)

    # Дублируем годовое значение на каждый месяц
    monthly_records = []
    for _, r in df.iterrows():
        for month in range(1, 13):
            date = f'{int(r["year"])}-{month:02d}'
            monthly_records.append({
                'date': date,
                'region': int(r['region']),
                'income_mln': r['income_mln']
            })

    df_monthly = pd.DataFrame(monthly_records).sort_values(['date', 'region']).reset_index(drop=True)
    df_monthly.to_csv(f'{OUT_DIR}/income.csv', index=False)

    print(f'income.csv: {len(df_monthly)} строк')
    print(f'Регионов сопоставлено: {df["region"].nunique()}')
    if unmatched:
        print(f'Не найдено ({len(unmatched)}):')
        for u in unmatched[:10]:
            print(f'  - {u}')


# ─────────────────────────────────────────────
# Запуск
# ─────────────────────────────────────────────
if __name__ == '__main__':
    print('=== Ключевая ставка + инфляция ===')
    process_cbr_inflation()

    print('\n=== Курс доллара ===')
    process_exchange_rate()

    print('\n=== Доходы населения по регионам ===')
    process_income()

    print('\nГотово! Файлы в data/macro/processed/')
